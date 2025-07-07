
"""
file_handler.py

此模块定义了 FileHandler 类，用于处理应用程序中的文件操作，
包括Excel文件的打开、保存、另存为，以及完整包（Excel文件和相关assets）的导入和导出。
"""

import os
import shutil
import zipfile
from datetime import datetime
from tkinter import filedialog, messagebox
import pandas as pd
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

from utils import Utils
from dependencies import PANDAS_AVAILABLE, XLSXWRITER_AVAILABLE, OPENPYXL_AVAILABLE

class FileHandler:
    """处理文件操作的类，包括Excel的导入导出和完整包的管理。"""

    def __init__(self, sheet, assets_dir, status_var):
        """
        初始化文件处理器。

        Args:
            sheet: tksheet 表格实例，用于获取和设置表格数据。
            assets_dir (str): 存储图片等资产的目录路径。
            status_var: Tkinter StringVar，用于更新状态栏信息。
        """
        self.sheet = sheet
        self.assets_dir = assets_dir
        self.status_var = status_var
        self.current_file = None  # 当前打开的Excel文件路径

        # 确保assets目录存在
        if not os.path.exists(self.assets_dir):
            os.makedirs(self.assets_dir)

    def save_excel_file(self):
        """保存Excel文件。如果已打开文件，则直接保存；否则另存为。"""
        if self.current_file:
            self._save_to_excel_file(self.current_file)
        else:
            self.save_as_excel_file()

    def save_as_excel_file(self):
        """另存为Excel文件。"""
        filename = filedialog.asksaveasfilename(
            title="保存Excel文件",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        
        if filename:
            self._save_to_excel_file(filename)
            self.current_file = filename

    def _save_to_excel_file(self, filename):
        """
        保存到Excel文件，包含图片和assets文件夹。
        将当前表格数据保存为Excel文件，并将assets目录中的图片复制到Excel文件同级目录下的assets文件夹。

        Args:
            filename (str): 要保存的Excel文件路径。
        """
        if not XLSXWRITER_AVAILABLE:
            messagebox.showerror("错误", "需要安装xlsxwriter库才能保存Excel文件。")
            return

        try:
            # 构建Excel文件对应的assets目录路径
            file_dir = os.path.dirname(filename)
            file_base = os.path.splitext(os.path.basename(filename))[0]
            excel_assets_dir = os.path.join(file_dir, f"{file_base}_assets")
            
            # 创建Excel文件对应的assets目录
            if not os.path.exists(excel_assets_dir):
                os.makedirs(excel_assets_dir)
            
            # 复制所有assets文件到Excel文件对应的assets目录
            if os.path.exists(self.assets_dir):
                for file in os.listdir(self.assets_dir):
                    src_path = os.path.join(self.assets_dir, file)
                    dst_path = os.path.join(excel_assets_dir, file)
                    if os.path.isfile(src_path):
                        shutil.copy2(src_path, dst_path)
            
            # 创建Excel工作簿和工作表
            workbook = xlsxwriter.Workbook(filename)
            worksheet = workbook.add_worksheet("表格数据")
            
            # 定义Excel单元格格式
            header_format = workbook.add_format({
                "bold": True,
                "bg_color": "#D7E4BC",
                "border": 1
            })
            
            cell_format = workbook.add_format({
                "border": 1,
                "text_wrap": True,
                "valign": "top"
            })
            
            # 写入表头
            headers = self.sheet.headers()
            for col, header in enumerate(headers):
                worksheet.write(0, col, str(header), header_format)
            
            # 写入数据（所有单元格都按字符串处理）
            data = self.sheet.get_sheet_data()
            for row, row_data in enumerate(data):
                for col, cell_value in enumerate(row_data):
                    if cell_value:
                        # 将所有内容都转换为字符串
                        cell_str = str(cell_value)
                        worksheet.write(row + 1, col, cell_str, cell_format)
                        
                        # 处理图片（插入到Excel中）
                        image_paths = Utils.extract_image_paths(cell_value)
                        for i, image_path in enumerate(image_paths):
                            # 使用Excel文件目录下的assets路径
                            excel_image_path = os.path.join(excel_assets_dir, os.path.basename(image_path))
                            if os.path.exists(excel_image_path):
                                try:
                                    worksheet.insert_image(row + 1, col, excel_image_path, {
                                        "x_scale": 0.3,
                                        "y_scale": 0.3,
                                        "x_offset": 5,
                                        "y_offset": 5 + i * 50
                                    })
                                except Exception as img_e:
                                    print(f"插入图片失败: {img_e}") # 打印错误但继续，不中断保存过程
            
            # 设置列宽
            for col in range(len(headers)):
                worksheet.set_column(col, col, 20)
            
            workbook.close()
            
            self.status_var.set(f"已保存到: {filename}")
            messagebox.showinfo("成功", f"Excel文件保存成功！\n图片文件已保存到: {excel_assets_dir}")
            
        except Exception as e:
            messagebox.showerror("错误", f"保存失败: {str(e)}")

    def open_excel_file(self):
        """打开Excel文件。"""
        filename = filedialog.askopenfilename(
            title="打开Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        
        if filename:
            self._load_from_excel_file(filename)

    def _load_from_excel_file(self, filename):
        """
        从Excel文件加载数据，并处理其中包含的图片。
        将Excel文件中的数据加载到tksheet表格中，并将其assets文件夹中的图片复制到应用程序的assets目录。

        Args:
            filename (str): 要加载的Excel文件路径。
        """
        if not PANDAS_AVAILABLE or not OPENPYXL_AVAILABLE:
            messagebox.showerror("错误", "需要安装pandas和openpyxl库才能打开Excel文件。")
            return

        try:
            # 使用pandas读取Excel文件，所有数据都按字符串处理
            df = pd.read_excel(filename, header=0, dtype=str)
            
            # 填充NaN值为空字符串，确保数据完整性
            df = df.fillna("")
            
            # 转换为列表格式，获取表头和数据
            headers = [str(col) for col in df.columns.tolist()]
            data = df.values.tolist()
            
            # 确保所有数据都是字符串类型
            for row in data:
                for i, cell in enumerate(row):
                    row[i] = str(cell) if cell is not None else ""
            
            # 确保有足够的行和列，以适应tksheet的初始大小或数据量
            min_rows = max(100, len(data))
            min_cols = max(20, len(headers))
            
            # 扩展数据到最小尺寸，如果原始数据小于最小尺寸，则用空字符串填充
            for row in data:
                while len(row) < min_cols:
                    row.append("")
            
            while len(data) < min_rows:
                data.append([""] * min_cols)
            
            # 生成Excel风格的列标题
            column_headers = Utils.generate_column_headers(min_cols)
            
            # 处理Excel文件目录下的assets文件夹
            file_dir = os.path.dirname(filename)
            file_base = os.path.splitext(os.path.basename(filename))[0]
            excel_assets_dir = os.path.join(file_dir, f"{file_base}_assets")
            
            # 如果存在Excel文件的assets目录，复制其内容到当前应用程序的assets目录
            if os.path.exists(excel_assets_dir):
                if not os.path.exists(self.assets_dir):
                    os.makedirs(self.assets_dir)
                
                for file in os.listdir(excel_assets_dir):
                    src_path = os.path.join(excel_assets_dir, file)
                    dst_path = os.path.join(self.assets_dir, file)
                    if os.path.isfile(src_path):
                        # 如果目标文件已存在，重命名以避免冲突
                        if os.path.exists(dst_path):
                            name, ext = os.path.splitext(file)
                            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                            new_name = f"{name}_{timestamp}{ext}"
                            dst_path = os.path.join(self.assets_dir, new_name)
                        
                        shutil.copy2(src_path, dst_path)
                
                self.status_var.set(f"已从 {excel_assets_dir} 导入图片文件")
            
            # 设置tksheet表格的数据和表头
            self.sheet.set_sheet_data(data)
            self.sheet.headers(column_headers)
            
            # 设置当前文件路径
            self.current_file = filename
            
            # 自动调整行高以适应内容
            Utils.auto_adjust_row_heights(self.sheet)
            
            self.status_var.set(f"已打开Excel文件: {filename}")
            messagebox.showinfo("成功", f"Excel文件打开成功！\n导入了 {len(data)} 行 {len(column_headers)} 列数据")
            
        except Exception as e:
            messagebox.showerror("错误", f"打开Excel文件失败: {str(e)}")

    def import_package(self):
        """导入完整包（ZIP文件包含Excel文件和assets文件夹）。"""
        filename = filedialog.askopenfilename(
            title="导入完整包",
            filetypes=[("ZIP文件", "*.zip"), ("所有文件", "*.*")]
        )
        
        if not filename:
            return
        
        try:
            import tempfile
            with tempfile.TemporaryDirectory() as temp_dir:
                # 解压ZIP文件到临时目录
                with zipfile.ZipFile(filename, "r") as zipf:
                    zipf.extractall(temp_dir)
                
                # 查找解压后的Excel文件
                excel_files = []
                for root, dirs, files in os.walk(temp_dir):
                    for file in files:
                        if file.endswith((".xlsx", ".xls")):
                            excel_files.append(os.path.join(root, file))
                
                if not excel_files:
                    messagebox.showerror("错误", "ZIP包中没有找到Excel文件")
                    return
                
                # 默认使用找到的第一个Excel文件
                excel_file = excel_files[0]
                
                # 查找解压后的assets文件夹
                assets_dirs = []
                for root, dirs, files in os.walk(temp_dir):
                    if "assets" in dirs:
                        assets_dirs.append(os.path.join(root, "assets"))
                
                # 复制assets文件夹内容到应用程序的assets目录
                if assets_dirs:
                    assets_src_dir = assets_dirs[0]
                    if not os.path.exists(self.assets_dir):
                        os.makedirs(self.assets_dir)
                    
                    for file in os.listdir(assets_src_dir):
                        src_path = os.path.join(assets_src_dir, file)
                        dst_path = os.path.join(self.assets_dir, file)
                        if os.path.isfile(src_path):
                            # 处理文件名冲突
                            if os.path.exists(dst_path):
                                name, ext = os.path.splitext(file)
                                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                                new_name = f"{name}_{timestamp}{ext}"
                                dst_path = os.path.join(self.assets_dir, new_name)
                            
                            shutil.copy2(src_path, dst_path)
                
                # 加载Excel文件到表格
                self._load_from_excel_file(excel_file)
                
                messagebox.showinfo("成功", "完整包导入成功！")
                self.status_var.set("完整包导入成功")
                
        except Exception as e:
            messagebox.showerror("错误", f"导入失败: {str(e)}")

    def export_package(self, wiki_exporter_instance):
        """
        导出完整包（ZIP文件包含Excel文件和Wiki文件）。

        Args:
            wiki_exporter_instance: WikiExporter 类的实例，用于获取Wiki内容。
        """
        filename = filedialog.asksaveasfilename(
            title="导出完整包",
            defaultextension=".zip",
            filetypes=[("ZIP文件", "*.zip"), ("所有文件", "*.*")]
        )
        
        if not filename:
            return
        
        try:
            with zipfile.ZipFile(filename, "w", zipfile.ZIP_DEFLATED) as zipf:
                # 添加assets文件夹中的所有文件到ZIP包
                if os.path.exists(self.assets_dir):
                    for root, dirs, files in os.walk(self.assets_dir):
                        for file in files:
                            file_path = os.path.join(root, file)
                            # arcname 是文件在ZIP包中的相对路径
                            arcname = os.path.relpath(file_path, ".")
                            zipf.write(file_path, arcname)
                
                # 导出Excel文件到临时位置并添加到ZIP包
                temp_excel = "temp_export.xlsx"
                self._save_excel_sync(temp_excel)
                zipf.write(temp_excel, "表格数据.xlsx")
                os.remove(temp_excel)
                
                # 导出Wiki文件到临时位置并添加到ZIP包
                temp_wiki = "temp_export.txt"
                wiki_exporter_instance._export_wiki_sync(temp_wiki) # 调用WikiExporter的同步导出方法
                zipf.write(temp_wiki, "表格数据_wiki.txt")
                os.remove(temp_wiki)
            
            messagebox.showinfo("成功", "完整包导出成功！")
            self.status_var.set("完整包导出成功")
            
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {str(e)}")

    def _save_excel_sync(self, filename):
        """
        同步保存Excel（用于打包）。
        此方法用于在不弹出文件对话框的情况下，将当前表格数据保存为Excel文件。

        Args:
            filename (str): 要保存的Excel文件路径。
        """
        workbook = xlsxwriter.Workbook(filename)
        worksheet = workbook.add_worksheet("表格数据")
        
        # 设置格式
        header_format = workbook.add_format({
            "bold": True,
            "bg_color": "#D7E4BC",
            "border": 1
        })
        
        cell_format = workbook.add_format({
            "border": 1,
            "text_wrap": True,
            "valign": "top"
        })
        
        # 写入表头
        headers = self.sheet.headers()
        for col, header in enumerate(headers):
            worksheet.write(0, col, str(header), header_format)
        
        # 写入数据（所有单元格都按字符串处理）
        data = self.sheet.get_sheet_data()
        for row, row_data in enumerate(data):
            for col, cell_value in enumerate(row_data):
                if cell_value:
                    # 将所有内容都转换为字符串
                    cell_str = str(cell_value)
                    worksheet.write(row + 1, col, cell_str, cell_format)
        
        # 设置列宽
        for col in range(len(headers)):
            worksheet.set_column(col, col, 20)
        
        workbook.close()


